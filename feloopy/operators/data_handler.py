# Copyright (c) 2022-2025, Keivan Tafakkori. All rights reserved.
# See the file LICENSE file for licensing details.

import numpy as np
import pandas as pd
import itertools as it
import os
import json

try:
    from ..extras.operators.data_handler import *
except:

    class FileManager:
        pass

class DataToolkit(FileManager):

    def __init__(self, key=None, memorize=True, measure=True):
        
        self.data = dict()
        self.seed= key
        self.random = np.random.default_rng(key)
        self.lfe = self.load_from_excel
        self.memorize=memorize
        self.gaussian = self.normal
        self.store = self.param =self.par = self.__keep
        self.measure = measure
        self.size = 0
        self.max_among_all_params = float('-inf')
        self.min_among_all_params = float('+inf')
        self.minimum_params = {}
        self.maximum_params = {}
        self.average_params = {}
        self.size_params = {}
        self.type_params = {}
        self.possible_epsilon = 1e-16
        self.possible_big_m = 1e16
        self.std_params = {}

    def sets(self,*args):
        return it.product(*args)
    
    def __fix_dims(self, dim, is_range=True):
        if dim == 0:
            pass
        elif not isinstance(dim, set):
            if len(dim) >= 1:
                if is_range:
                    dim = [range(d) if not isinstance(d, range) else d for d in dim]
                else:
                    dim = [len(d) if not isinstance(d, int) else d for d in dim]
        return dim

    def __calculate_total_size(self, data):
        if isinstance(data, (list, tuple)):
            return sum(self.__calculate_total_size(item) for item in data)
        elif isinstance(data,set):
            return 0
        elif isinstance(data, dict):
            return sum(self.__calculate_total_size(key) + self.__calculate_total_size(value) for key, value in data.items())
        elif isinstance(data, np.ndarray):
            return data.size
        elif isinstance(data, pd.DataFrame):
            return data.size
        elif isinstance(data, pd.Series):
            return data.size
        elif hasattr(data, '__len__') and not isinstance(data, (str, bytes)):
            return sum(self.__calculate_total_size(item) for item in data)
        else:
            return 1

    def __calculate_stats(self, data):
        def flatten(data):
            if isinstance(data, (list, tuple)):
                for item in data:
                    yield from flatten(item)
            elif isinstance(data, set):
                for item in data:
                    yield from flatten(item)
            elif isinstance(data, dict):
                for key, value in data.items():
                    yield from flatten(key)
                    yield from flatten(value)
            elif isinstance(data, np.ndarray):
                for item in data.flatten():
                    yield from flatten(item)
            elif isinstance(data, pd.Series):
                for item in data.values.flatten():
                    yield from flatten(item)
            elif isinstance(data, pd.DataFrame):
                for item in data.values.flatten():
                    yield from flatten(item)
            elif hasattr(data, '__iter__') and not isinstance(data, (str, bytes)):
                for item in data:
                    yield from flatten(item)
            elif isinstance(data, (int, float, str, bytes, complex, bool)):
                yield data
            elif isinstance(data, np.generic):
                yield data.item()
            else:
                yield data

        values = list(flatten(data))

        if not values:
            return float('inf'), float('-inf'), float('nan'), float('nan')

        if not all(isinstance(x, (int, float)) or isinstance(x, type(values[0])) for x in values):
            return '-', len(values), None, None, None, None
        

        type_checks = {
            'ℝ': all(isinstance(x, (int, float)) for x in values) and any(x > 0 for x in values) and any(x < 0 for x in values), 
            'ℝ⁺': all(isinstance(x, (int, float)) and x >= 0 for x in values) and any(x > 0 for x in values) and any(x > 1 for x in values),
            'ℝ⁻': all(isinstance(x, (int, float)) and x <= 0 for x in values) and any(x < 0 for x in values) and any(x < -1 for x in values), 
            'ℤ': all(isinstance(x, int) for x in values) and any(x > 0 for x in values) and any(x < 0 for x in values),
            'ℤ⁺': all(isinstance(x, int) and x >= 0 for x in values) and any(x > 0 for x in values), 
            'ℕ': all(isinstance(x, int) and x > 0 for x in values),
            'ℤ₀': all(isinstance(x, int) and x == 0 for x in values), 
            'ℤ⁻': all(isinstance(x, int) and x < 0 for x in values),  
            '𝔹': all(isinstance(x, int) and x in [0, 1] for x in values) and any(x == 1 for x in values),  
            'ℝ⁺ ∩ [0, 1]': all(isinstance(x, (int,float)) and 0 <= x <= 1 for x in values) and any(0 < x < 1 for x in values), 
            'ℝ⁻ ∩ [-1, 0]': all(isinstance(x, (int,float)) and -1 <= x <= 0 for x in values) and any(-1 < x < 0 for x in values),  
            'Σ': all(isinstance(x, str) for x in values), 
            'B': all(isinstance(x, bytes) for x in values), 
            'ℂ': all(isinstance(x, complex) for x in values), 
            '∅': all(x is None for x in values), 
            '-': not all(isinstance(x, (int, float,str,bytes,complex)) for x in values), 
        }

        identified_types = [name for name, check in type_checks.items() if check]
        data_type = '/'.join([t.strip() for t in identified_types]) if identified_types else 'Other'
        data_type+="    "

        if 'ℝ' in data_type or 'ℤ' in data_type: 
            numeric_values = [x for x in values if isinstance(x, (int, float))]
            if numeric_values:
                min_value = min(numeric_values)
                max_value = max(numeric_values)
                mean_value = sum(numeric_values) / len(numeric_values)
                std_deviation = (sum((x - mean_value) ** 2 for x in numeric_values) / len(numeric_values)) ** 0.5
            else:
                min_value = max_value = mean_value = std_deviation = "-       "
        else:
            numeric_values = [x for x in values if isinstance(x, (int, float))]
            min_value = min(numeric_values) if numeric_values else "-       "
            max_value = max(numeric_values) if numeric_values else "-       "
            mean_value = std_deviation = "-       "

        data_size = len(values)
        return data_type, data_size, min_value, max_value, mean_value, std_deviation

    def __keep(self, name, value, neglect=False):
        if self.measure == True:
            self.type_params[name], self.size_params[name], self.minimum_params[name],self.maximum_params[name],self.average_params[name],self.std_params[name] = self.__calculate_stats(value)
            try:
                self.max_among_all_params = max(self.maximum_params[name],self.max_among_all_params)
                self.min_among_all_params = min(self.minimum_params[name],self.min_among_all_params)
            except:
                pass
            try:
                self.possible_epsilon  = 1/self.max_among_all_params
            except:
                pass
            self.possible_big_m = self.max_among_all_params
            try:
                self.size+=self.__calculate_total_size(value)
            except:
                print("warning: exception for {name} in size calculation. Ignoring real size.")
                self.size+=1
        if self.memorize and neglect==False:
            self.data[name]=value
            return self.data[name]
        elif neglect:
            return value
        else:
            return value
    
    # === Sets

    def _convert_to_set(self, input_set):
        if isinstance(input_set, set):
            return input_set
        elif isinstance(input_set, range):
            return set(input_set)
        elif isinstance(input_set, list):
            return set(input_set)
        else:
            raise TypeError("Unsupported set type")

    def _json_serializer(self, obj):
        if isinstance(obj, np.ndarray):
            return {
                "__type__": "ndarray",
                "shape": obj.shape,
                "data": obj.tolist()
            }
        elif isinstance(obj, pd.DataFrame):
            return {
                "__type__": "dataframe",
                "index": obj.index.tolist(),
                "columns": obj.columns.tolist(),
                "data": obj.to_dict(orient='records')
            }
        elif isinstance(obj, pd.Series):
            return {
                "__type__": "series",
                "index": obj.index.tolist(),
                "data": obj.to_dict()
            }
        elif isinstance(obj, set):
            return {
                "__type__": "set",
                "data": list(obj)
            }
        elif isinstance(obj, range):
            return {
                "__type__": "range",
                "start": obj.start,
                "stop": obj.stop,
                "step": obj.step
            }
        raise TypeError(f"Object of type {type(obj).__name__} is not JSON serializable")

    def _json_decoder(self, dct):
        if "shape" in dct and "data" in dct:
            return np.array(dct["data"]).reshape(dct["shape"])
        elif "columns" in dct and "index" in dct and "data" in dct:
            return pd.DataFrame(data=dct["data"], index=dct["index"], columns=dct["columns"])
        elif "index" in dct and "data" in dct:
            try:
                return pd.Series(data=dct["data"], index=dct["index"])
            except Exception:
                pass  
        elif isinstance(dct, dict) and "data" in dct:
            if "__type__" in dct and dct["__type__"] == "set":
                return set(dct["data"])
        elif "start" in dct and "stop" in dct and "step" in dct:
            return range(dct["start"], dct["stop"], dct["step"])
        return dct

    def set(
        self,
        name,
        bound=None,
        step=1,
        callback=None,
        to_list=False,
        to_range=False,
        named_indices=False,
        size=None,
        init=None,
        axis=0,
        neglect=False):
        
        if size is not None:
            if to_range:
                result = range(size)
            else:
                result = set(range(size))
        elif init is not None:
            
            if type(init)==int:
                if to_range:
                    result = range(init)
                else:
                    result = set(range(init))
            elif type(init)==np.ndarray:
                if to_range:
                    result = range(np.shape(init)[axis])
                else:
                    result = set(range(np.shape(init)[axis]))

            elif type(init)==list:
                if to_range:
                    result = range(len(list))
                else:
                    result = set(range(len(list)))
            else:
                try:
                    result = set(init)
                except:
                    result = init 

            if callback:
                result =  set(item for item in result if callback(item))

        else:   
            if callback:
                named_indices = False
               
            if to_range:
                result = range(bound[0], bound[1] + 1, step)
            elif named_indices:
                result = {f"{name.lower()}{i}" for i in range(bound[0], bound[1] + 1, step) if not callback or callback(i)}
            else:
                if callback:
                    result =  set(item for item in range(bound[0], bound[1] + 1, step) if callback(item))
                else:
                    result =  set(range(bound[0], bound[1] + 1, step))
        if to_list:
            result = list(result)
        
        return self.__keep(name, result, neglect)
    
    def array(self, input):
        return np.array(input)

    def alias(self, name, init, neglect=False):
        result=init
        return self.__keep(name, result, neglect)

    def union(self, name, *sets, neglect=False):
        converted_sets = [self._convert_to_set(s) for s in sets]
        result=  set().union(*converted_sets)
        return self.__keep(name, result, neglect)

    def intersection(self, name, *sets, neglect=False):
        converted_sets = [self._convert_to_set(s) for s in sets]
        result = set().intersection(*converted_sets)
        return self.__keep(name, result, neglect)

    def difference(self,name, *sets, neglect=False):
        converted_sets = [self._convert_to_set(s) for s in sets]
        result = converted_sets[0]
        for s in converted_sets[1:]:
            result = result.difference(s)
        return self.__keep(name, result, neglect)

    def symmetric_difference(self,name, *sets, neglect=False):
        converted_sets = [self._convert_to_set(s) for s in sets]
        result = converted_sets[0]
        for s in converted_sets[1:]:
            result = result.symmetric_difference(s)
        return self.__keep(name, result, neglect)

    def display(self, input=None):
        from pprint import pprint
        if input:
            pprint(input, width=90, indent=0, sort_dicts=True)
        else:
            pprint(self.data, width=90, indent=0, sort_dicts=True)
    
    # === Parameters

    def _sample_list_or_array(self, name, init, size, replace=False, sort_result=False, return_indices=False, axis=None):
        
        if isinstance(init, (list, range)):
            init = np.array(init)

        if axis is None:
            sampled_indices = self.random.choice(init.size, size=size, replace=replace)
        else:
            axis = np.atleast_1d(axis)
            
            for ax in axis:
                axis_size = init.shape[ax]
                sampled_indices = self.random.choice(axis_size, size=size, replace=replace)
            
                init = np.take(init, sampled_indices, axis=ax)
        
        if return_indices:
            if sort_result:
                sampled_indices.sort()
            result = sampled_indices
        else:
            result = init
            if sort_result:
                result.sort()

        return result

    def _sample_pandas_dataframe(self, name, init, size, replace=False, sort_result=False, return_indices=False, axis=None):
        axis = 0 if axis is None else axis 

        if axis not in [0, 1]:
            raise ValueError("Invalid axis for Pandas DataFrame sampling. Supported axes: 0 (rows), 1 (columns)")

        sampled_indices = self.random.choice(init.shape[axis], size=size, replace=replace)

        if return_indices:
            sampled_data = sampled_indices
        else:
            if axis == 0:
                sampled_data = init.iloc[sampled_indices, :]
            else:
                sampled_data = init.iloc[:, sampled_indices]

            if sort_result:
                sampled_data = sampled_data.sort_index(axis=axis)
                
        return sampled_data

    def zeros(self, name, dim=0, neglect=False):
        dim = self.__fix_dims(dim,is_range=False)
        if dim == 0:
            result = np.zeros(1)
        else:
            if type(dim)==set:
                result = {key: 0 for key in dim}
            else:
                result = np.zeros(dim)
        return self.__keep(name, result, neglect)

    def ones(self, name, dim=0, neglect=False):
        dim = self.__fix_dims(dim,is_range=False)
        if dim == 0:
            result = np.ones(1)
        else:
            if type(dim)==set:
                result = {key: 0 for key in dim}
            else:
                result = np.ones(tuple(dim))
        return self.__keep(name, result, neglect)

    def ones_per_column(self, name, dim=0, min_ones=1, max_ones=1, neglect=False):
        dim = self.__fix_dims(dim,is_range=False)
        if dim == 0:
            result = np.ones(1)
        else:
            rows, cols = dim
            if type(rows)!=int:
                rows = len(rows)
            if type(cols)!=int:
                cols = len(cols)
            result = np.zeros((rows, cols))
            for col in range(cols):
                num_ones = self.random.integers(min_ones, max_ones + 1)
                rows_with_ones = self.random.choice(rows, num_ones, replace=False)
                result[rows_with_ones, col] = 1
        if type(dim)==set:
            result = {key: result[key] for key in dim}

        return self.__keep(name, result, neglect)
    
    def ones_per_row(self, name, dim=0, min_ones=1, max_ones=1, neglect=False):
        dim = self.__fix_dims(dim,is_range=False)
        if dim == 0:
            result = np.ones(1)
        else:
            rows, cols = dim
            if type(rows)!=int:
                rows = len(rows)
            if type(cols)!=int:
                cols = len(cols)
            result = np.zeros((rows, cols))
            for row in range(rows):
                num_ones = self.random.integers(min_ones, max_ones + 1)
                cols_with_ones = self.random.choice(cols, num_ones, replace=False)
                result[row, cols_with_ones] = 1
        if type(dim)==set:
            result = {key: result[key] for key in dim}

        return self.__keep(name, result, neglect)

    def permutation(self, name, dim=0, neglect=False):
        dim = self.__fix_dims(dim,is_range=False)
        if len(dim)!=2:
            raise ValueError("Only 2D matrices.")
        if dim[0]!=dim[1]:
            raise ValueError("Only box matrices (i.e., n=m)")

        identity_matrix = np.eye(dim[0])
        self.random.shuffle(identity_matrix)
        result = identity_matrix
        if type(dim)==set:
            result = {key: result[key] for key in dim}     
            
        return self.__keep(name, result, neglect)
    
    def uniformint(self, name, dim=0, bound=[1, 10], neglect=False):
        dim = self.__fix_dims(dim,is_range=False)
        if dim == 0:
            result = self.random.integers(low=bound[0], high=bound[1] + 1)
        else:
            if type(dim)==set:
                result = {key: self.random.integers(low=bound[0], high=bound[1] + 1) for key in dim}
            else:
                result = self.random.integers(low=bound[0], high=bound[1] + 1, size=dim)
        return self.__keep(name, result, neglect)
    
    def bernoulli(self, name, dim=0, p=0.5, neglect=False):
        dim = self.__fix_dims(dim,is_range=False)
        if dim == 0:
            result = self.random.choice([0, 1], p=[1-p, p])
        else:
            if type(dim)==set:
                result = {key: self.random.choice([0, 1], p=[1-p, p]) for key in dim}
            else:
                result = self.random.choice([0, 1], p=[1-p, p], size=dim)
        return self.__keep(name, result, neglect)
    
    def binomial(self, name, dim=0, n=None, p=None, neglect=False):
        dim = self.__fix_dims(dim, is_range=False)
        if dim == 0:
            result = self.random.binomial(n, p)
        else:
            if type(dim)==set:
                result = {key: self.random.binomial(n, p) for key in dim}
            else:
                result = self.random.binomial(n, p, size=tuple(dim))
        return self.__keep(name, result, neglect)

    def poisson(self, name, dim=0, lam=1, neglect=False):
        dim = self.__fix_dims(dim, is_range=False)
        if dim == 0:
            result = self.random.poisson(lam)
        else:
            if type(dim)==set:
                result = {key: self.random.poisson(lam) for key in dim}
            else:
                result = self.random.poisson(lam, size=tuple(dim))
        return self.__keep(name, result, neglect)

    def geometric(self, name, dim=0, p=None, neglect=False):
        dim = self.__fix_dims(dim, is_range=False)
        if dim == 0:
            result = self.random.geometric(p)
        else:
            if type(dim)==set:
                result = {key: self.random.geometric(p) for key in dim}
            else: 
                result = self.random.geometric(p, size=tuple(dim))
        return self.__keep(name, result, neglect)

    def negative_binomial(self, name, dim=0, r=None, p=None, neglect=False):
        dim = self.__fix_dims(dim, is_range=False)
        if dim == 0:
            result = self.random.negative_binomial(r, p)
        else:
            if type(dim)==set:
                result = {key: self.random.negative_binomial(r, p) for key in dim}
            else:
                result = self.random.negative_binomial(r, p, size=tuple(dim))
        return self.__keep(name, result, neglect)

    def hypergeometric(self, name, dim=0, N=None, m=None, n=None, neglect=False):
        nbad = m
        ngood = N - m
        nsamples = n

        dim = self.__fix_dims(dim, is_range=False)
        if dim == 0:
            result = self.random.hypergeometric(ngood, nbad, nsamples)
        else:
            if type(dim)==set:
                result = {key: self.random.hypergeometric(ngood, nbad, nsamples) for key in dim}
            else:
                result = self.random.hypergeometric(ngood, nbad, nsamples, size=tuple(dim))
        return self.__keep(name, result, neglect)

    def uniform(self, name, dim=0, bound=[0, 1], neglect=False):
        dim = self.__fix_dims(dim,is_range=False)
        if dim == 0:
            result = self.random.uniform(low=bound[0], high=bound[1])
        else:
            if type(dim)==set:
                result = {key: self.random.uniform(low=bound[0], high=bound[1]) for key in dim}
            else:
                result = self.random.uniform(low=bound[0], high=bound[1], size=dim)
        return self.__keep(name, result, neglect)

    def rsum(self, name, total: float, dim=0, bound=[0, 1], neglect=False):
        if total <= 0:
            raise ValueError(f"Target sum must be positive. Got: {total}")

        dim = self.__fix_dims(dim, is_range=False)
        if isinstance(dim, set):
            dim = len(dim)

        if dim == 0:
            val = self.random.uniform(low=bound[0], high=bound[1])
            return self.__keep(name, total if not neglect else val, neglect)

        try:
            data = self.random.uniform(low=bound[0], high=bound[1], size=dim)
            s = data.sum()
            if s == 0:
                raise ValueError("Generated random values sum to zero, cannot normalize")
            scaled = data * (total / s)
        except Exception as e:
            raise RuntimeError(f"Failed to generate scaled random values: {e}")

        return self.__keep(name, scaled, neglect)

    def normal(self, name, dim=0, mu=0, sigma=1, neglect=False):
        dim = self.__fix_dims(dim,is_range=False)
        if dim == 0:
            result = self.random.normal(mu, sigma)
        else:
            if type(dim)==set:
                result = {key: self.random.normal(mu, sigma) for key in dim}
            else:
                result = self.random.normal(mu, sigma, size=dim)
        return self.__keep(name, result, neglect)

    def standard_normal(self, name, dim=0, neglect=False):
        dim = self.__fix_dims(dim,is_range=False)
        if dim == 0:
            result = self.random.normal(0, 1)
        else:
            if type(dim)==set:
                result = {key: self.random.normal(0, 1) for key in dim}
            else:
                result = self.random.normal(0, 1, size=dim)
        return self.__keep(name, result, neglect)

    def exponential(self, name, dim=0, lam=1.0, neglect=False):
        dim = self.__fix_dims(dim, is_range=False)
        if dim == 0:
            result = self.random.exponential(scale=1/lam)
        else:
            if type(dim)==set:
                result = {key: self.random.exponential(scale=1/lam) for key in dim}
            else:
                result = self.random.exponential(scale=1/lam, size=dim)
        return self.__keep(name, result, neglect)

    def gamma(self, name, dim=0, alpha=1, lam=1, neglect=False):
        dim = self.__fix_dims(dim, is_range=False)
        if dim == 0:
            result = self.random.gamma(shape=alpha, scale=1/lam)
        else:
            if type(dim)==set:
                result = {key: self.random.gamma(shape=alpha, scale=1/lam) for key in dim}
            else:
                result = self.random.gamma(shape=alpha, scale=1/lam, size=dim)
        return self.__keep(name, result, neglect)

    def erlang(self, name, dim=0, alpha=1, lam=1, neglect=False):
        alpha = int(alpha)
        dim = self.__fix_dims(dim, is_range=False)
        if dim == 0:
            result = self.random.gamma(shape=alpha, scale=1/lam)
        else:
            if type(dim)==set:
                result = {key: self.random.gamma(shape=alpha, scale=1/lam) for key in dim}
            else:
                result = self.random.gamma(shape=alpha, scale=1/lam, size=dim)
        return self.__keep(name, result, neglect)

    def beta(self, name, dim=0, a=1, b=1, neglect=False):
        dim = self.__fix_dims(dim, is_range=False)
        if dim == 0:
            result = self.random.beta(a, b, size=None)
        else:
            if type(dim)==set:
                result = {key: self.random.beta(a, b)  for key in dim}
            else:
                result = self.random.beta(a, b, size=dim)
        return self.__keep(name, result, neglect)

    def weibull(self, name, dim=0, alpha=None, beta=None, neglect=False):
        dim = self.__fix_dims(dim, is_range=False)
        if dim == 0:
            result = alpha * self.random.weibull(a=beta)
        else:
            if type(dim)==set:
                result = {key: alpha * self.random.weibull(a=beta) for key in dim}
            else: 
                result = alpha * self.random.weibull(a=beta, size=dim)
        return self.__keep(name, result, neglect)

    def cauchy(self, name, dim=0, neglect=False):
        dim = self.__fix_dims(dim, is_range=False)
        if dim == 0:
            result = self.random.standard_cauchy()
        else:
            if type(dim)==set:
                result = {key: self.random.standard_cauchy() for key in dim}
            else:
                result = self.random.standard_cauchy(size=dim)
        return self.__keep(name, result, neglect)

    def dirichlet(self, name, dim=0, k=None, alpha=None, neglect=False):
        dim = self.__fix_dims(dim, is_range=False)
        if alpha is None:
            if k is not None:
                alpha = np.ones(k)
            elif isinstance(dim, list):
                alpha = np.ones(len(dim[-1]))
        if dim == 0 or len(dim) == 1:
            result = self.random.dirichlet(alpha)
        else:
            if type(dim)==set:
                result = {key: self.random.dirichlet(alpha) for key in dim}
            else:
                result = self.random.dirichlet(alpha, size=dim)
        return self.__keep(name, result, neglect)

    def colors(self, name, dim=0, neglect=False, with_names=True):
        import matplotlib.colors as mcolors
        colors_dict = dict(mcolors.BASE_COLORS, **mcolors.CSS4_COLORS)
        dim = self.__fix_dims(dim,is_range=False)
        dim = [range(i) for i in dim]
        if dim == 0:    
            if with_names:
                result = self.random.choice(list(colors_dict.keys()))
            else:
                result = '#{:06x}'.format(self.random.integers(0, 0xFFFFFF))
        else:
            if len(dim) == 1:
                if with_names:
                    result = {key: self.random.choice(list(colors_dict.keys())) for key in dim[0]}
                else:
                    result = {key: '#{:06x}'.format(self.random.integers(0, 0xFFFFFF)) for key in dim[0]}
            else:
                if with_names:
                    result = {key: self.random.choice(list(colors_dict.keys())) for key in it.product(*dim)}
                else:
                    result = {key: '#{:06x}'.format(self.random.integers(0, 0xFFFFFF)) for key in it.product(*dim)}
        return self.__keep(name, result, neglect)
    
    def distance(self, name, dim=0, bound=[0, 1], symmetric=True, as_int=False, neglect=False):
        dim = self.__fix_dims(dim, is_range=False)
        if as_int:
            mat = self.random.integers(low=bound[0], high=bound[1] + 1, size=dim)
        else:
            mat = self.random.uniform(low=bound[0], high=bound[1], size=dim)

        if symmetric:
            mat = (mat + mat.T) / 2
            if as_int:
                mat = np.round(mat).astype(int)
        np.fill_diagonal(mat, 0)
        return self.__keep(name, mat, neglect)
    
    def points(
        self,
        name,
        n=0,
        dim=2,
        bound=None,
        geo=False,
        as_int=False,
        country=None,
        city=None,
        custom_polygon=None,
        mode="euclidean",
        max_travel_time=None,
        max_tries=5,
        return_distances=False,
        neglect=False
    ):
        try:
            import numpy as np
        except ImportError:
            raise ImportError("NumPy is required. Install via `pip install numpy`.")

        if not isinstance(n, int) or n < 0:
            raise ValueError(f"`n` must be a nonnegative integer; got {n}.")
        if not isinstance(dim, int) or dim < 0:
            raise ValueError(f"`dim` must be a nonnegative integer; got {dim}.")

        allowed_modes = {"walk", "drive", "ship", "train", "air", "euclidean", "manhattan"}
        if mode not in allowed_modes:
            raise ValueError(f"`mode` must be one of {allowed_modes}; got '{mode}'.")

        speed_map = {
            "walk": 5.0,
            "drive": 60.0,
            "ship": 30.0,
            "train": 80.0,
            "air": 800.0
        }

        def _haversine_matrix(latlons):
            R = 6371.0
            lat = np.radians(latlons[:, 0])[:, None]
            lon = np.radians(latlons[:, 1])[:, None]
            dlat = lat - lat.T
            dlon = lon - lon.T
            a = (np.sin(dlat / 2))**2 + np.cos(lat) * np.cos(lat.T) * (np.sin(dlon / 2))**2
            c = 2 * np.arctan2(np.sqrt(a), np.sqrt(1 - a))
            return R * c

        def _sample_in_shape(n_pts, shape, as_int_flag):
            try:
                from shapely.geometry import Point
            except ImportError:
                raise ImportError("Shapely is required. Install via `pip install shapely`.")
            rng = self.random
            minx, miny, maxx, maxy = shape.bounds
            pts = []
            batch = max(n_pts * 3, 100)
            while len(pts) < n_pts:
                xs = rng.uniform(minx, maxx, size=batch)
                ys = rng.uniform(miny, maxy, size=batch)
                for x, y in zip(xs, ys):
                    p = Point(x, y)
                    if shape.contains(p):
                        pts.append((y, x))
                        if len(pts) == n_pts:
                            break
            arr = np.array(pts, dtype=float)
            if as_int_flag:
                arr = np.round(arr).astype(int)
            return arr

        if not hasattr(self, "_graph_cache"):
            self._graph_cache = {}
        if not hasattr(self, "_countries_gdf"):
            self._countries_gdf = None

        cache_dir = "osm_cache"
        os.makedirs(cache_dir, exist_ok=True)

        def _network_samples_and_distances(n_pts, city_name, net_mode, as_int_flag):
            try:
                import osmnx as ox
            except ImportError:
                raise ImportError("OSMnx is required. Install via `pip install osmnx`.")
            try:
                import networkx as nx
            except ImportError:
                raise ImportError("NetworkX is required. Install via `pip install networkx`.")

            safe_city = city_name.replace(" ", "_").replace(",", "")
            filename = f"{safe_city}_{net_mode}.graphml"
            filepath = os.path.join(cache_dir, filename)

            if (city_name, net_mode) in self._graph_cache:
                G = self._graph_cache[(city_name, net_mode)]
            elif os.path.isfile(filepath):
                try:
                    G = ox.load_graphml(filepath)
                except Exception:
                    G = ox.graph_from_place(city_name, network_type=net_mode if net_mode in ("walk", "drive") else "all",
                                            custom_filter='["route"="ferry"]' if net_mode == "ship" else None)
                    ox.save_graphml(G, filepath)
                self._graph_cache[(city_name, net_mode)] = G
            else:
                if net_mode in ("walk", "drive"):
                    G = ox.graph_from_place(city_name, network_type=net_mode)
                else:
                    G = ox.graph_from_place(
                        city_name,
                        network_type="all",
                        custom_filter='["route"="ferry"]'
                    )
                    if len(G.nodes) == 0:
                        raise ValueError(f"No ferry/ship routes found in '{city_name}'.")
                ox.save_graphml(G, filepath)
                self._graph_cache[(city_name, net_mode)] = G

            rng = self.random
            all_nodes = list(G.nodes)
            if n_pts > len(all_nodes):
                raise ValueError(f"Requested {n_pts} points, but graph has only {len(all_nodes)} nodes.")
            chosen_nodes = rng.choice(all_nodes, size=n_pts, replace=False)

            latlons = []
            for node in chosen_nodes:
                data = G.nodes[node]
                lat = data.get("y")
                lon = data.get("x")
                if lat is None or lon is None:
                    raise RuntimeError(f"Node {node} lacks 'x'/'y'.")
                latlons.append((lat, lon))
            pts_arr = np.array(latlons, dtype=float)
            if as_int_flag:
                pts_arr = np.round(pts_arr).astype(int)

            dist_mat = np.zeros((n_pts, n_pts), dtype=float)
            for i, src in enumerate(chosen_nodes):
                lengths = nx.single_source_dijkstra_path_length(G, src, weight="length")
                for j, tgt in enumerate(chosen_nodes):
                    dist_mat[i, j] = lengths.get(tgt, np.inf)
            return pts_arr, dist_mat / 1000.0

        def _geo_samples_and_haversine(n_pts, as_int_flag):
            if city is not None:
                try:
                    import osmnx as ox
                except ImportError:
                    raise ImportError("OSMnx is required. Install via `pip install osmnx`.")
                try:
                    gdf_c = ox.geocode_to_gdf(city)
                except Exception as e:
                    raise ValueError(f"OSMnx could not geocode '{city}': {e}")
                if gdf_c.empty:
                    raise ValueError(f"City '{city}' not found by OSMnx.")
                shape = gdf_c.unary_union
                return _sample_in_shape(n_pts, shape, as_int_flag)
            if country is not None:
                try:
                    import geopandas as gpd
                except ImportError:
                    raise ImportError("GeoPandas is required. Install via `pip install geopandas`.")
                if self._countries_gdf is None:
                    try:
                        self._countries_gdf = gpd.read_file(
                            "https://raw.githubusercontent.com/datasets/geo-countries/master/data/countries.geojson"
                        )
                    except Exception as e:
                        raise RuntimeError(f"Failed to load country boundaries. Details: {e}")
                world = self._countries_gdf
                cols = world.columns
                for candidate in ("ADMIN", "admin", "NAME", "name", "Country", "country"):
                    if candidate in cols:
                        name_col = candidate
                        break
                else:
                    raise ValueError(f"No country-name column found. Columns: {cols.tolist()}")
                matches = world[world[name_col].str.lower() == country.lower()]
                if matches.empty:
                    sample_names = ", ".join(sorted(world[name_col].unique())[:10]) + ", …"
                    raise ValueError(f"Country '{country}' not found. Examples: {sample_names}")
                shape = matches.iloc[0].geometry
                return _sample_in_shape(n_pts, shape, as_int_flag)
            if custom_polygon is not None:
                try:
                    from shapely.geometry import Polygon, MultiPolygon
                except ImportError:
                    raise ImportError("Shapely is required. Install via `pip install shapely`.")
                if dim != 2:
                    raise ValueError("custom_polygon requires dim=2.")
                if not isinstance(custom_polygon, (Polygon, MultiPolygon)):
                    raise ValueError("custom_polygon must be a Shapely Polygon or MultiPolygon.")
                return _sample_in_shape(n_pts, custom_polygon, as_int_flag)
            if geo:
                if bound is None:
                    bound_list = [[-90, 90], [-180, 180]]
                else:
                    if (
                        isinstance(bound, (list, tuple))
                        and len(bound) == 2
                        and all(np.isscalar(x) for x in bound)
                    ):
                        bound_list = [list(bound) for _ in range(2)]
                    elif (
                        isinstance(bound, (list, tuple))
                        and len(bound) == 2
                        and all(isinstance(axis, (list, tuple)) and len(axis) == 2 for axis in bound)
                    ):
                        bound_list = [list(axis) for axis in bound]
                    else:
                        raise ValueError("`bound` for geo sampling must be [lat_min, lat_max], [lon_min, lon_max].")
                lat_min, lat_max = bound_list[0]
                lon_min, lon_max = bound_list[1]
                if not (-90 <= lat_min <= 90 and -90 <= lat_max <= 90):
                    raise ValueError(f"Latitude bounds must lie in [-90,90]. Got [{lat_min},{lat_max}].")
                if not (-180 <= lon_min <= 180 and -180 <= lon_max <= 180):
                    raise ValueError(f"Longitude bounds must lie in [-180,180]. Got [{lon_min},{lon_max}].")
                lats = self.random.uniform(lat_min, lat_max, size=n_pts)
                lons = self.random.uniform(lon_min, lon_max, size=n_pts)
                pts_arr = np.column_stack((lats, lons))
                if as_int_flag:
                    pts_arr = np.round(pts_arr).astype(int)
                return pts_arr
            raise ValueError("Geo sampling requires city, country, custom_polygon, or geo=True.")

        if mode in ("walk", "drive", "ship"):
            if dim != 2:
                raise ValueError(f"mode='{mode}' only valid when dim=2.")
            speed = speed_map[mode]
            for attempt in range(max_tries):
                if city:
                    pts_arr, dist_mat = _network_samples_and_distances(n, city, mode, as_int)
                else:
                    pts_arr = _geo_samples_and_haversine(n, as_int)
                    dist_mat = _haversine_matrix(pts_arr)
                if max_travel_time is None:
                    break
                travel_time = dist_mat / speed
                if np.all(np.isfinite(travel_time) & (travel_time <= max_travel_time)):
                    break
            else:
                raise ValueError(
                    f"Could not generate {n} points within {max_travel_time}h by {mode} after {max_tries} tries."
                )
            stored = self.__keep(name, pts_arr, neglect)
            dist_mat = self.__keep(name+"_dist", dist_mat, neglect)
            if return_distances:
                return stored, dist_mat
            return stored

        if mode in ("train", "air"):
            if dim != 2:
                raise ValueError(f"mode='{mode}' requires dim=2.")
            speed = speed_map[mode]
            for attempt in range(max_tries):
                pts_arr = _geo_samples_and_haversine(n, as_int)
                if max_travel_time is not None:
                    dist_mat = _haversine_matrix(pts_arr)
                    travel_time = dist_mat / speed
                    if np.all(np.isfinite(travel_time) & (travel_time <= max_travel_time)):
                        break
                else:
                    break
            else:
                raise ValueError(
                    f"Could not generate {n} points within {max_travel_time}h by {mode} after {max_tries} tries."
                )
            stored = self.__keep(name, pts_arr, neglect)
            if return_distances:
                dist_mat = _haversine_matrix(pts_arr)
                dist_mat = self.__keep(name+"_dist", dist_mat, neglect)
                return stored, dist_mat
            return stored

        if mode in ("euclidean", "manhattan"):
            if bound is None:
                bound_list = [[0, 1]] * dim
            else:
                if (
                    isinstance(bound, (list, tuple))
                    and len(bound) == 2
                    and all(np.isscalar(x) for x in bound)
                ):
                    bound_list = [list(bound) for _ in range(dim)]
                elif (
                    isinstance(bound, (list, tuple))
                    and len(bound) == dim
                    and all(isinstance(axis, (list, tuple)) and len(axis) == 2 for axis in bound)
                ):
                    bound_list = [list(axis) for axis in bound]
                else:
                    raise ValueError(f"`bound` must be [low, high] or list of length {dim} of [low_i, high_i].")
            for i, (low_i, high_i) in enumerate(bound_list):
                if not (np.isscalar(low_i) and np.isscalar(high_i)):
                    raise ValueError(f"`bound[{i}]` values must be numeric; got {bound_list[i]}.")
                if low_i > high_i:
                    raise ValueError(f"`bound[{i}] = [{low_i},{high_i}]` invalid: min > max.")
            for attempt in range(max_tries):
                pts_arr = np.empty((n, dim), dtype=float)
                for i in range(dim):
                    low_i, high_i = bound_list[i]
                    pts_arr[:, i] = self.random.uniform(low_i, high_i, size=n)
                if as_int:
                    pts_arr = np.round(pts_arr).astype(int)
                if max_travel_time is not None:
                    if mode == "euclidean":
                        diff = pts_arr[:, None, :] - pts_arr[None, :, :]
                        dist_mat = np.sqrt((diff**2).sum(axis=2))
                    else:
                        diff = pts_arr[:, None, :] - pts_arr[None, :, :]
                        dist_mat = np.abs(diff).sum(axis=2)
                    travel_time = dist_mat
                    if np.all(np.isfinite(travel_time) & (travel_time <= max_travel_time)):
                        break
                else:
                    break
            else:
                raise ValueError(
                    f"Could not generate {n} points within {max_travel_time}h by {mode} after {max_tries} tries."
                )
            stored = self.__keep(name, pts_arr, neglect)
            if return_distances:
                if mode == "euclidean":
                    diff = pts_arr[:, None, :] - pts_arr[None, :, :]
                    dist_mat = np.sqrt((diff**2).sum(axis=2))
                else:
                    diff = pts_arr[:, None, :] - pts_arr[None, :, :]
                    dist_mat = np.abs(diff).sum(axis=2)
                    dist_mat = self.__keep(name+"_dist", dist_mat, neglect)
                return stored, dist_mat
            return stored

        raise RuntimeError(f"Unhandled mode '{mode}'.")

    def sample(self, name, init, size, replace=False, sort_result=False, reset_index=False, return_indices=False, axis=None, neglect=False):

        type_is= type(init)
        if type_is ==set:
            init = list(init)
            
        if isinstance(init, (list, set,range, np.ndarray)):
            sample =  self._sample_list_or_array(name, init, size, replace, sort_result, return_indices, axis)
        elif isinstance(init, pd.DataFrame):
            sample = self._sample_pandas_dataframe(name, init, size, replace, sort_result, return_indices, axis)
        else:
            raise ValueError("Unsupported data type for sampling. Supported types: set, list, range, numpy.ndarray, pandas.DataFrame")

        if reset_index:
            sample = sample.reset_index(drop=True)

        if type_is ==set:
            sample =  set(sample)
        elif type_is in [list, range]:
            sample = list(sample)
        elif return_indices:
            sample =  set(sample)
        else:
            sample =  sample

        return self.__keep(name, sample, neglect)

    def set_local_parameters(self):

        for key, value in self.data.items():
            locals()[key] = value      

    def set_global_parameters(self):

        for key, value in self.data.items():
            globals()[key] = value

    def load_from_excel(
        self,
        name: str,
        dim=0,
        labels: list = None,
        appearance: list = None,
        file_name: str = "data.xlsx",
        neglect: bool = False
    ):

        
        if labels is None and type(dim)!=int:
            labels=["" for d in dim]

        if dim==0:
            labels=[""]
            appearance=[0,0]

        if isinstance(dim, list) and len(dim) >= 1 and isinstance(dim[0], set):
            dim = [len(d) for d in dim]
        dim = self.__fix_dims(dim, is_range=True)

        def _format_index_key(level_vals, label, key):
            if label:
                return label + str(key)
            if key in level_vals:
                return key
            ks = str(key)
            if ks in level_vals:
                return ks
            return key

        if len(appearance) == 2:
            if (
                (appearance[0] == 1 and appearance[1] == 1)
                or (appearance[0] == 1 and appearance[1] == 0)
                or (appearance[0] == 0 and appearance[1] == 0)
                or (appearance[0] == 0 and appearance[1] == 1)
            ):
                result = pd.read_excel(
                    file_name, index_col=0, sheet_name=name
                ).to_numpy()

            else:
                header_arg = [i for i in range(appearance[1])] if appearance[1] > 0 else None
                index_arg = [i for i in range(appearance[0])] if appearance[0] > 0 else None

                try:
                    parameter = pd.read_excel(
                        file_name,
                        header=header_arg,
                        index_col=index_arg,
                        sheet_name=name,
                    )
                except Exception as e:
                    raise ValueError(
                        f"Cannot read sheet '{name}' with header={header_arg} "
                        f"and index_col={index_arg}:\n  {e}"
                    )

                created_par = np.zeros(tuple(len(x) for x in dim), dtype=float)

                row_index = parameter.index
                col_index = parameter.columns

                for keys in it.product(*dim):
                    try:
                        row_elems = []
                        if isinstance(row_index, pd.MultiIndex):
                            for i in range(appearance[0]):
                                level_vals = row_index.levels[i]
                                row_elems.append(_format_index_key(level_vals, labels[i], keys[i]))
                        else:
                            level_vals = row_index
                            row_elems.append(_format_index_key(level_vals, labels[0], keys[0]))

                        row_key = tuple(row_elems) if appearance[0] > 1 else row_elems[0]
                        col_elems = []
                        if appearance[1] > 0:
                            if isinstance(col_index, pd.MultiIndex):
                                for j in range(appearance[1]):
                                    level_vals = col_index.levels[j]
                                    col_elems.append(
                                        _format_index_key(
                                            level_vals,
                                            labels[appearance[0] + j],
                                            keys[appearance[0] + j],
                                        )
                                    )
                                col_key = tuple(col_elems)
                            else:
                                level_vals = col_index
                                col_elems.append(
                                    _format_index_key(
                                        level_vals,
                                        labels[appearance[0]],
                                        keys[appearance[0]],
                                    )
                                )
                                col_key = col_elems[0]
                        else:
                            col_key = None
                        if appearance[0] == 0:
                            val = parameter.loc[:, col_key]
                        elif appearance[1] == 0:
                            val = parameter.loc[row_key]
                        else:
                            val = parameter.loc[row_key, col_key]
                        if isinstance(val, (pd.Series, pd.DataFrame, np.ndarray)):
                            arr = np.array(val).flatten()
                            val = arr[0] if arr.size > 0 else np.nan
                        created_par[keys] = val
                    except Exception:
                        created_par[keys] = np.nan

                result = created_par

        else:
            par = pd.read_excel(file_name, index_col=0, sheet_name=name).to_numpy()
            result = par.reshape(par.shape[0],)

        if dim == 0:
            result = result[0][0]
        elif len(dim) == 1:
            result = np.reshape(result, [len(dim[0]),])
        else:
            pass

        return self.__keep(name, result, neglect)
    
    def save_to_excel(
        self,
        name: str,
        array: np.ndarray,
        dim=0,
        labels: list = None,
        appearance: list = None,
        file_name: str = "data.xlsx"
    ) -> None:
        import pandas as pd
        from openpyxl import Workbook, load_workbook
        import itertools as it
        import os
        import numpy as np

        if not (isinstance(dim, (list, tuple)) or isinstance(dim, int)):
            raise TypeError("dim must be int or list/tuple")

        if appearance is None:
            if isinstance(dim, int):
                appearance = [0, max(1, array.ndim - 1)]
            else:
                appearance = [0, max(1, len(dim) - 1)]

        if labels is None:
            if isinstance(dim, int):
                labels = [""]
            else:
                labels = [""] * len(dim)

        if isinstance(dim, (list, tuple)) and dim and isinstance(dim[0], set):
            dim = [len(s) for s in dim]

        if hasattr(self, '__fix_dims'):
            try:
                dim = self.__fix_dims(dim, is_range=True)
            except Exception as e:
                raise RuntimeError(f"Error in __fix_dims: {e}")

        def _size(x):
            return x if isinstance(x, int) else len(x)

        if not (isinstance(appearance, (list, tuple)) and len(appearance) == 2):
            raise ValueError("appearance must be a list or tuple of length 2")

        nr, nc = appearance

        if isinstance(dim, int):
            dim_len = 1
        else:
            dim_len = len(dim)

        if not (isinstance(nr, int) and isinstance(nc, int)):
            raise TypeError("appearance values must be integers")

        if nr < 0 or nc < 0 or nr + nc > dim_len:
            raise ValueError(f"Invalid appearance values: nr={nr}, nc={nc}, dim length={dim_len}")

        if labels is not None and len(labels) < dim_len:
            raise ValueError(f"labels length {len(labels)} is less than dim length {dim_len}")

        arr = np.asarray(array, dtype=float)

        try:
            shape = tuple(_size(d) for d in dim) if dim else arr.shape
        except Exception as e:
            raise ValueError(f"Error processing dim values: {e}")

        try:
            mat = arr.reshape(shape)
        except Exception as e:
            raise ValueError(f"Cannot reshape array of size {arr.size} into shape {shape}: {e}")

        col_dims = dim[nr:nr + nc]
        total_columns = 1
        for d in col_dims:
            total_columns *= _size(d)

        MAX_COLUMNS = 16384

        try:
            if total_columns <= MAX_COLUMNS:
                wb = load_workbook(file_name) if os.path.exists(file_name) else Workbook()
                if name in wb.sheetnames:
                    std = wb[name]
                    wb.remove(std)
                ws = wb.create_sheet(title=name)

                dim_sizes = [_size(d) for d in col_dims]
                all_col_keys = list(it.product(*(range(s) for s in dim_sizes)))

                for level in range(nc):
                    for col_idx, keys in enumerate(all_col_keys):
                        label_val = f"{labels[nr + level]}{keys[level]}"
                        ws.cell(row=level + 1, column=nr + col_idx + 1, value=label_val)

                row_dims = dim[:nr]
                row_keys = list(it.product(*(range(_size(d)) for d in row_dims))) if nr else [()]

                for row_idx, rk in enumerate(row_keys):
                    for i in range(nr):
                        ws.cell(row=nc + 1 + row_idx, column=i + 1, value=f"{labels[i]}{rk[i]}")

                for row_idx, rk in enumerate(row_keys):
                    for col_idx, ck in enumerate(all_col_keys):
                        full_idx = rk + ck
                        try:
                            val = mat[full_idx]
                        except Exception as e:
                            raise IndexError(f"Failed to access element at index {full_idx}: {e}")
                        ws.cell(row=nc + 1 + row_idx, column=nr + col_idx + 1, value=float(val))

                if 'Sheet' in wb.sheetnames and len(wb.sheetnames) > 1:
                    std = wb['Sheet']
                    wb.remove(std)

                wb.save(file_name)
                wb.close()
                return

        except Exception as e:
            print(f"Warning: Couldn't save in 2D format ({e}). Falling back to long format.")

        dim_names = labels[:len(dim)] if labels else [f"dim_{i}" for i in range(len(dim))]
        all_indices = list(it.product(*(range(_size(d)) for d in dim)))
        data = []
        for idx in all_indices:
            try:
                val = float(arr[idx])
            except Exception as e:
                raise IndexError(f"Error accessing array element {idx}: {e}")
            data.append((*idx, val))

        df = pd.DataFrame(data, columns=[*dim_names, "Value"])

        if os.path.exists(file_name):
            try:
                book = load_workbook(file_name)
                if name in book.sheetnames:
                    std = book[name]
                    book.remove(std)
                book.save(file_name)
                book.close()
            except Exception as e:
                print(f"Warning: Couldn't modify existing file ({e}). Overwriting.")

        try:
            with pd.ExcelWriter(
                file_name,
                engine='openpyxl',
                mode='a' if os.path.exists(file_name) else 'w'
            ) as writer:
                df.to_excel(writer, sheet_name=name, index=False)
        except Exception as e:
            raise IOError(f"Failed to write dataframe to Excel: {e}")

    def save(self, name, format="json"):
        directory = os.path.join('results', 'data')
        if not os.path.exists(directory):
            os.makedirs(directory)

        extension = format.lower()
        file_path = os.path.abspath(os.path.join(directory, f"{name}.{extension}"))

        if format == "json":
            try:
                with open(file_path, 'w') as file:
                    json.dump(self.data, file, default=self._json_serializer, indent=4)
                print(f"Data successfully exported to JSON at: {file_path}")
            except TypeError as e:
                print(f"Serialization error: {e}")
            except Exception as e:
                print(f"An error occurred while saving JSON: {e}")

        elif format == "parquet":
            try:
                df = pd.DataFrame(self.data)
                df.to_parquet(file_path, index=False)
                print(f"Data successfully exported to Parquet at: {file_path}")
            except Exception as e:
                print(f"An error occurred while saving Parquet: {e}")

        else:
            print(f"Unsupported format: {format}. Supported formats: 'json', 'parquet'")

    def load(self, name, format="json", neglect=False):
        extension = format.lower()
        filename = f"{name}.{extension}"
        final_dir = os.path.join('.', 'data', 'final')
        results_dir = os.path.join('results', 'data')
        file_path_final = os.path.join(final_dir, filename)
        file_path_results = os.path.join(results_dir, filename)

        if os.path.exists(file_path_final):
            file_path = os.path.abspath(file_path_final)
        elif os.path.exists(file_path_results):
            file_path = os.path.abspath(file_path_results)
            print(f"Data file found in {results_dir}. For consistency, please move it to {final_dir}.")
        else:
            raise FileNotFoundError(f"File '{filename}' not found. Please place it in '{final_dir}' or '{results_dir}'.")

        data = None
        try:
            if extension == "json":
                with open(file_path, 'r') as file:
                    data = json.load(file, object_hook=self._json_decoder)
                print(f"Data successfully imported from JSON at: {file_path}")

            elif extension == "parquet":
                df = pd.read_parquet(file_path)
                data = df.to_dict(orient='records')
                print(f"Data successfully imported from Parquet at: {file_path}")

            else:
                raise ValueError(f"Unsupported format: {format}. Supported formats are 'json' and 'parquet'.")
        except Exception as e:
            print(f"An error occurred while loading {format.upper()}: {e}")
            data = None

        return self.__keep(name, data, neglect)
    
data_toolkit = DataToolkit
